"""
upload_referrals.py
=====================
엑셀 두 파일을 읽어 Firebase Firestore에 업로드합니다.

■ db_import_ready.xlsx  (251행)
  - referrals 컬렉션용 데이터 없음 (referrer_wallet_address 전부 None)
  - users 컬렉션에 지갑 + 성과 메타 저장

■ 5_6228882217737658062.xlsx  (443행)
  - 컬럼: id | 用户地址(본인) | 上级地址(추천인) | ...
  - referrals 컬렉션에 referrerWallet / referredWallet 쌍 저장

Firestore 컬렉션 매핑
  referrals : {referrerWallet, referredWallet, referrerCode, createdAt, updatedAt}
  referral_users : {walletAddress, referralCount, totalTeamMembers, personalPerformance,
                    totalTeamPerformance, totalNodeValue, investmentProducts, dailyIncome, ...}
"""

import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore
import time
import sys

# ── Firebase 초기화 (API Key 방식 → REST fallback) ───────────────────────────
# firebase-admin은 서비스 계정이 필요하지만, 여기서는
# python-firebase REST API를 직접 사용합니다.

import requests, json, os

PROJECT_ID = os.environ.get("FIREBASE_PROJECT_ID", "numine-dev-e4ec1")
API_KEY    = os.environ.get("FIREBASE_API_KEY", "")
BASE_URL   = f"https://firestore.googleapis.com/v1/projects/{PROJECT_ID}/databases/(default)/documents"

def fs_write(collection: str, doc_id: str, fields: dict) -> dict:
    """Firestore REST API로 단일 문서 생성/덮어쓰기"""
    url = f"{BASE_URL}/{collection}/{doc_id}?key={API_KEY}"
    # field 타입 자동 변환
    def to_fv(v):
        if v is None:                return {"nullValue": None}
        if isinstance(v, bool):      return {"booleanValue": v}
        if isinstance(v, int):       return {"integerValue": str(v)}
        if isinstance(v, float):     return {"doubleValue": v}
        if isinstance(v, str):       return {"stringValue": v}
        return {"stringValue": str(v)}
    payload = {"fields": {k: to_fv(v) for k, v in fields.items()}}
    r = requests.patch(url, json=payload, timeout=10)
    return r.json()

def fs_exists(collection: str, field: str, value: str) -> bool:
    """특정 필드값으로 문서 존재 여부 확인 (structuredQuery)"""
    url = f"https://firestore.googleapis.com/v1/projects/{PROJECT_ID}/databases/(default)/documents:runQuery?key={API_KEY}"
    body = {
        "structuredQuery": {
            "from": [{"collectionId": collection}],
            "where": {
                "fieldFilter": {
                    "field": {"fieldPath": field},
                    "op": "EQUAL",
                    "value": {"stringValue": value}
                }
            },
            "limit": 1
        }
    }
    r = requests.post(url, json=body, timeout=10)
    docs = r.json()
    return bool(docs and docs[0].get("document"))

# ── 파일 1: db_import_ready.xlsx → referral_users 컬렉션 ──────────────────
print("=" * 60)
print("파일 1: db_import_ready.xlsx → referral_users 컬렉션")
print("=" * 60)

wb1 = openpyxl.load_workbook("/home/user/uploaded_files/db_import_ready.xlsx")
ws1 = wb1["Sheet1"]

ok1 = 0
skip1 = 0
err1 = 0

for r in range(2, ws1.max_row + 1):
    wallet  = ws1.cell(r, 1).value
    if not wallet:
        continue
    wallet = str(wallet).strip().lower()

    referral_code       = ws1.cell(r, 2).value or ""
    status              = ws1.cell(r, 3).value or ""
    joined_at_raw       = ws1.cell(r, 4).value
    personal_perf       = float(ws1.cell(r, 5).value or 0)
    total_team_perf     = float(ws1.cell(r, 6).value or 0)
    total_team_members  = int(ws1.cell(r, 7).value or 0)
    total_node_value    = float(ws1.cell(r, 8).value or 0)
    investment_products = float(ws1.cell(r, 9).value or 0)
    daily_income        = float(ws1.cell(r, 10).value or 0)
    referrer_wallet     = ws1.cell(r, 11).value or ""

    # joined_at: datetime → ms
    if hasattr(joined_at_raw, "timestamp"):
        joined_at = int(joined_at_raw.timestamp() * 1000)
    else:
        joined_at = int(time.time() * 1000)

    doc_id = wallet
    fields = {
        "walletAddress":       wallet,
        "referralCode":        str(referral_code),
        "status":              str(status),
        "joinedAt":            joined_at,
        "personalPerformance": personal_perf,
        "totalTeamPerformance": total_team_perf,
        "totalTeamMembers":    total_team_members,
        "totalNodeValue":      total_node_value,
        "investmentProducts":  investment_products,
        "dailyIncome":         daily_income,
        "referrerWallet":      str(referrer_wallet).lower() if referrer_wallet else "",
        "updatedAt":           int(time.time() * 1000),
    }

    res = fs_write("referral_users", doc_id, fields)
    if "error" in res:
        print(f"  ✗ [{r}] {wallet[:16]}… → {res['error'].get('message','?')}")
        err1 += 1
    else:
        ok1 += 1
        if ok1 % 50 == 0:
            print(f"  ✓ {ok1}건 완료...")

print(f"\n✅ referral_users: 성공 {ok1}건 / 오류 {err1}건\n")

# ── 파일 2: 5_6228882217737658062.xlsx → referrals 컬렉션 ──────────────────
print("=" * 60)
print("파일 2: 5_6228882217737658062.xlsx → referrals 컬렉션")
print("=" * 60)

wb2 = openpyxl.load_workbook("/home/user/uploaded_files/5_6228882217737658062.xlsx")
ws2 = wb2["Sheet1"]

ok2 = 0
skip2 = 0
err2 = 0

now_ms = int(time.time() * 1000)

for r in range(2, ws2.max_row + 1):
    row_id      = ws2.cell(r, 1).value
    referred    = ws2.cell(r, 2).value   # 用户地址 (본인)
    referrer    = ws2.cell(r, 3).value   # 上级地址 (추천인)

    if not referred:
        continue

    referred = str(referred).strip().lower()
    referrer = str(referrer).strip().lower() if referrer else ""

    # 루트 노드(추천인 없음) 또는 더미 주소 스킵
    if not referrer or referrer == "0x8888888888888888888888888888888888888888":
        skip2 += 1
        continue

    # 문서 ID: referrer_referred (중복 방지)
    doc_id = f"{referrer}_{referred}"

    fields = {
        "referrerWallet": referrer,
        "referredWallet": referred,
        "referrerCode":   "",          # 원본 데이터에 코드 없음
        "createdAt":      now_ms,
        "updatedAt":      now_ms,
    }

    res = fs_write("referrals", doc_id, fields)
    if "error" in res:
        print(f"  ✗ [{r}] {referred[:14]}… → {res['error'].get('message','?')}")
        err2 += 1
    else:
        ok2 += 1
        if ok2 % 50 == 0:
            print(f"  ✓ {ok2}건 완료...")

print(f"\n✅ referrals: 성공 {ok2}건 / 스킵(루트) {skip2}건 / 오류 {err2}건\n")
print("=" * 60)
print(f"전체 완료  referral_users {ok1}건 + referrals {ok2}건")
print("=" * 60)
